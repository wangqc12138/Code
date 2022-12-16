# import os
import openpyxl


def help(src: str, tar: str):
    dirs = []
    for line in open('./src.txt', 'r'):
        dirs.append(line.strip())
    # ex = openpyxl.load_workbook(src)
    # sh = ex['Sheet1']
    # map = dict()
    # for i in range(1, sh.max_column+1):
    #     if "/data" in sh.cell(row=1, column=i):
    #         map[sh.cell(row=2, column=i)] = sh.cell(row=1, column=i)
    for i in dirs:
        print(i)
    ex = openpyxl.load_workbook(tar)
    sh1 = ex['Sheet1']
    sh2 = ex['Sheet2']
    t = 1
    for dir in dirs:
        for j in range(2, 4):
            for i in range(0, sh1.max_column):
                ch = chr(i+ord('A'))
                if ch == 'L':
                    sh2[f'{ch}{t}'].value = sh1[f'{ch}{j}'].value.replace(
                        "XXX", dir)
                else:
                    sh2[f'{ch}{t}'].value = sh1[f'{ch}{j}'].value
            t += 1
    ex.save("res.xlsx")


help("./src.txt", "./test.xlsx")
