from collections import defaultdict
from ftplib import FTP
import ctypes
import os
import openpyxl

# def ftpcheck(host="", port=21, user="", passwd="", dir="", matchruler=""):
#     ftp = FTP()
#     ftp.connect(host=host, port=port, timeout=120)
#     ftp.login(user=user, passwd=passwd)
#     res = []
#     ftp.dir(dir+"/"+matchruler+"*", res.append)
#     index = 0
#     for line in res:
#         if line[0] != 'd':
#             index += 1
#     data = host+"|"+dir+"|"+matchruler+"|"+str(index)+"\n"
#     # file = open(log)
#     # file.write(data)
#     print(data)
#     return data


# def test():
#     str = "xferPasswd kt4"
#     passwd = os.popen(str)
#     print(passwd)
#     print(type(passwd))
#     passwd.readlines()
#     for line in passwd:
#         print(line)
# temp = os.popen(["xferPasswd -d ", passwd])
# print(temp)

def test1():
    wb = openpyxl.load_workbook('./文件大小（202204-202304).xlsx')
    all_sheet = wb.get_sheet_names()
    for s_name in all_sheet:
        sheet = wb.get_sheet_by_name(s_name)
        all_data = defaultdict(int)
        day = defaultdict(int)
        print(s_name)
        for i in range(1, sheet.max_row + 1):
            if isinstance(sheet.cell(i, 2).value, int):
                # print(sheet.cell(i, 1).value, sheet.cell(i, 2).value)
                # print(str(sheet.cell(i, 1).value)[0:6])
                all_data[str(sheet.cell(i, 1).value)[0:6]
                         ] += sheet.cell(i, 2).value
                day[str(sheet.cell(i, 1).value)[0:6]] += 1
        pre = -1
        temp = 0
        mon = 0
        t = 0
        for (x, y) in all_data.items():
            p = round(y/day[x], 2)
            if pre == -1:
                temp = p
                pre = p
            # print(x, p, '%.2f' % (p/pre), '%.2f' % (p/temp))
            t += p/pre
            pre = p

            mon += 1
        #     t = x
        print('%.2f' % (t/mon))
        # time = int(t[0:4])*12+int(t[5:6])
        # k = round(temp/mon, 2)
        # print(k)
        # for i in range(1, 21):
        #     st = str((time+i)//12)+str((time+i) % 12)
        #     pre *= k
        # print(st, pre)


# def save2db(data = ""):
#     return
if __name__ == '__main__':
    # data = ftpcheck("10.10.10.183", 21, "wangqc", "wangqichao",
    #                 "/data03/ktgrp/wangqc/dest", "111")
    # save2db(data)
    test1()
