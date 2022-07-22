import openpyxl
import os
import re
from copy import copy

def copyLowOneToSever(startLowCol: str, endLowCol:str, copyTime: int):
    lowLength = eval(re.findall(r'\d+', endLowCol)[0])
    colLength = ord(re.findall(r'[a-zA-Z]+', endLowCol)[0]) -  ord(re.findall(r'[a-zA-Z]+', startLowCol)[0]) + 1
    wb = openpyxl.load_workbook('模板.xlsx')
    wsCopyed = wb['Sheet3'] #工作簿
    wsProduction = wb['Sheet1']
    print(lowLength,colLength)
    for time in range(0 ,copyTime ):
        for m in range(1 ,  lowLength + 1):
            wsProduction.row_dimensions[m].height = wsCopyed.row_dimensions[m].height
            for n in range(1 ,  colLength+ 1):
                if n < 27:
                    c = chr(n + 64).upper()  
                else:
                    if n < 677:
                        c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                    else:
                        c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
                            divmod(divmod(n, 676)[1], 26)[1] + 64)

                i = '%s%d' % (c , m+ (lowLength*time))
                init = '%s%d' % (c, m)# 单元格编号
                if m == 1:
                    #				 print("Modify column %s width from %d to %d" % (n, ws2.column_dimensions[c].width ,ws.column_dimensions[c].width))
                    wsProduction.column_dimensions[c].width = wsCopyed.column_dimensions[c].width
                try:
                    cell1 = wsCopyed[init]  # 获取data单元格数据
                    wsProduction[i].value = cell1.value  # 赋值到ws2单元格
                    if cell1.has_style:  # 拷贝格式
                        wsProduction[i].font = copy(cell1.font)
                        wsProduction[i].border = copy(cell1.border)
                        wsProduction[i].fill = copy(cell1.fill)
                        wsProduction[i].number_format = copy(cell1.number_format)
                        wsProduction[i].protection = copy(cell1.protection)
                        wsProduction[i].alignment = copy(cell1.alignment)
                except AttributeError as e:
                    print("cell(%s) is %s" % (i, e))
                    continue
    wb.save('test1.xlsx')
    wb.close()

#copyLowOneToSever('A2','K104',52)
wb = openpyxl.load_workbook('./模板(1).xlsx')
sheet = wb['Sheet3']
f = open('./src.txt','r',encoding='utf-8')
row = 2
mylist = []
index = 0
rw=2
merge_lists = sheet.merged_cells
for i in merge_lists:
    r1, r2, c1, c2 = i.min_row, i.max_row, i.min_col, i.max_col
    if c1 == 4 and r1 != 1:
        mylist.append(r1)
mylist.sort()
for i in range(1,52):
    for j in range(0,23):
        mylist.append(mylist[j]+i*103)
for line in f:
    for i in range(2,105):
        sheet.cell(row,6).value = sheet.cell(i,6).value.replace('诺西端局M14语音',line.strip())
        sheet.cell(row,7).value = sheet.cell(i,7).value
        sheet.cell(row,8).value = sheet.cell(i,8).value.replace('诺西端局M14语音',line.strip())
        sheet.cell(row,9).value = sheet.cell(i,9).value.replace('诺西端局M14语音',line.strip())
        sheet.cell(row,10).value = sheet.cell(i,10).value
        row+=1
    for j in range(0,23):
        r = mylist[index]
        sheet.cell(r,4).value = sheet.cell(mylist[j],4).value.replace('诺西端局M14语音',line.strip())
        sheet.cell(r,5).value = sheet.cell(mylist[j],5).value.replace('诺西端局M14语音',line.strip())
        if index > 0:
            sheet.merge_cells(start_row=mylist[index-1],start_column=4,end_row=r-1,end_column=4)
            sheet.merge_cells(start_row=mylist[index-1],start_column=5,end_row=r-1,end_column=5)
        index+=1
    sheet.cell(rw,1).value = sheet.cell(2,3).value
    sheet.cell(rw,2).value = sheet.cell(2,2).value.replace('诺西端局M14语音',line.strip())
    sheet.cell(rw,3).value = sheet.cell(2,3).value.replace('诺西端局M14语音',line.strip())
    if rw > 103 :
        sheet.merge_cells(start_row=rw-103,start_column=1,end_row=rw-1,end_column=1)
        sheet.merge_cells(start_row=rw-103,start_column=2,end_row=rw-1,end_column=2)
        sheet.merge_cells(start_row=rw-103,start_column=3,end_row=rw-1,end_column=3)
    rw+=103
wb.save('result.xlsx')
